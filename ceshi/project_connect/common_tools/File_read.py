# -*- coding: utf-8 -*-
# @Time    : 2021/11/23 10:10
# @Author  : Hello world!
# @E-mail  : loading....
# @File    : File_read.py

from openpyxl import Workbook, load_workbook
import re

class Excel_object:
    def read_excel(self, sheet):
        rx = load_workbook(r'E:\pycharm\python\ceshi\project_connect\test_datas\test.xlsx')
        sheet_read = rx[sheet]
        title = []
        dict_datas = []
        dict_excel = {}
        for i in range(1, sheet_read.max_column+1):
            title.append(sheet_read.cell(1, i).value)
        for h in range(0, sheet_read.max_row-1):
            for j in range(2, sheet_read.max_column+2):
                dict_excel[title[j-2]] = sheet_read.cell(h+2, j-1).value
            dict_datas.append(dict_excel)
            dict_excel = {}
        return dict_datas

    def write_excel(self, data, position):
        wd = load_workbook(r'E:\pycharm\python\ceshi\project_connect\test_datas\test.xlsx')
        # wd.create_sheet('new_sheet_1', index=2)
        # wd.save(r'E:\pycharm\python\ceshi\project_connect\test_datas\test.xlsx')
        sheet_write = wd['数据']
        if str(data).isdigit():
            sheet_write.cell(2, int(position), int(data))
            sheet_write.cell(2, int(position)+1, int(data)+1)
            wd.save(r'E:\pycharm\python\ceshi\project_connect\test_datas\test.xlsx')
        else:
            find = re.findall(r'\d*\d', str(data))
            b = int(find[0]) + 1
            datas_new = data.replace(str(find[0]), str(b))
            sheet_write.cell(2, int(position), data)
            sheet_write.cell(2, int(position)+1, datas_new)
            wd.save(r'E:\pycharm\python\ceshi\project_connect\test_datas\test.xlsx')



# Excel_object().write_excel('发布会3', 5)
# a = Excel_object().read_excel('数据')
# print(a)


# a = Excel_object().read_excel('添加嘉宾')
# b = Excel_object().read_excel('数据')
# if a[0]['请求参数'].find('${phone}') != -1:
#     a[0]['请求参数'] = a[0]['请求参数'].replace('${phone}', str(b[0]['phone_new']))
# print(a[0]['请求参数'])

# def write_excel(self,row,coulmn,data):
#     wb=load_workbook(self.filename)
#     st=wb[self.sheetname]
#     st.cell(row,coulmn,data)
#     wb.save(self.filename)
#     pass
#
# if data["请求参数"].find("${new_eid}") != -1:
#     a=Do_excel(self.filename,"增加发布会1").read_excel()
#     data["请求参数"] = data["请求参数"].replace("${new_eid}", str(a[0]["new_eid"]))
#     print(data["请求参数"])
# res=Http_request().http_request(data["请求方式"],data["请求地址"],eval(data["请求参数"]))
# try:
#     a=res.json()["status"]
#     b=eval(data["请求参数"])["eid"]
#     Do_excel(self.filename,"增加发布会1").write_excel(2,1,b)
#     Do_excel(self.filename, "增加发布会1").write_excel(2, 2, b+1)
#     self.assertEqual(a,data["预期结果"],"测试失败")

#
# a = '啊1'
# b = list(a)
# c = int(b[-1])+1
# b[-1] = str(c)
# d = ''.join(b)
# print(d)

# wd = load_workbook(r'E:\pycharm\python\ceshi\project_connect\test_datas\test.xlsx')
# datas = list(data)
# b = int(datas[-1])+1
# datas[-1] = str(b)
# datas_new = ''.join(datas)
# sheet_write.cell(2, int(position), data)
# sheet_write.cell(2, int(position)+1, datas_new)
# wd.save(r'E:\pycharm\python\ceshi\project_connect\test_datas\test.xlsx')
